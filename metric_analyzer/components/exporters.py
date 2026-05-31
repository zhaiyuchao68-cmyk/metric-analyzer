"""导出功能：Excel 和 PDF 报告生成"""

import io
import base64
from pathlib import Path

import pandas as pd
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from metric_analyzer.models import DecompositionResult, MetricConfig, AnalysisMode, DecompositionMethod
from metric_analyzer.interpreters.nl_generator import NLGenerator
from metric_analyzer.components.charts import waterfall_chart, contribution_bar, composition_pie, dual_factor_grouped_bar


def export_excel(
    result: DecompositionResult,
    config: MetricConfig,
    original_data: pd.DataFrame,
    summary: str,
) -> bytes:
    """导出完整 Excel 报告

    Args:
        result: 分析结果
        config: 指标配置
        original_data: 原始数据
        summary: 自然语言解读

    Returns:
        Excel 文件的字节内容
    """
    wb = Workbook()

    # Sheet1: 原始数据
    ws_data = wb.active
    ws_data.title = "原始数据"
    _write_dataframe(ws_data, original_data, "原始数据")

    # Sheet2: 拆解结果
    ws_result = wb.create_sheet("拆解结果")
    _write_decomposition_result(ws_result, result, config)

    # Sheet3: 自然语言解读
    ws_summary = wb.create_sheet("分析解读")
    _write_summary(ws_summary, summary, config)

    # 保存到字节流
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _get_column_letter(col_idx: int) -> str:
    """获取列字母，避免 merged cell 问题"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _write_dataframe(ws, df: pd.DataFrame, title: str):
    """将 DataFrame 写入工作表"""
    from openpyxl.utils import get_column_letter

    # 标题行
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.cell(1, 1).font = Font(bold=True, size=14)

    # 空行
    ws.append([])

    # 表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(3, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, row in enumerate(df.values, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    # 自动调整列宽
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[col_letter].width = min(int(max_length) + 2, 30)


def _write_decomposition_result(ws, result: DecompositionResult, config: MetricConfig):
    """写入拆解结果"""
    # 标题
    ws.append([f"指标拆解结果：{config.name}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(1, 1).font = Font(bold=True, size=14)

    # 基本信息
    ws.append([])
    ws.append(["拆解方法", result.method.value])
    ws.append(["MECE 检验", "是" if result.is_mece else "否"])
    ws.append(["因子数量", len(result.contributions)])
    ws.append([])

    # 结果表格
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # 表头
    headers = ["因子", "贡献值", "贡献率(%)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(7, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, factor in enumerate(result.contributions, 8):
        ws.cell(row_idx, 1, factor.name)
        ws.cell(row_idx, 2, round(factor.value_change, 2))
        ws.cell(row_idx, 3, round(factor.contribution_rate * 100, 2))

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def _write_summary(ws, summary: str, config: MetricConfig):
    """写入自然语言解读"""
    ws.append([f"分析解读：{config.name}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.cell(1, 1).font = Font(bold=True, size=14)

    ws.append([])

    # 按段落写入
    for line in summary.split('\n'):
        if line.strip():
            ws.append([line])

    ws.column_dimensions['A'].width = 80


def export_pdf(
    result: DecompositionResult,
    config: MetricConfig,
    summary: str,
    top_n: int = 5,
) -> bytes:
    """导出 PDF 报告

    Args:
        result: 分析结果
        config: 指标配置
        summary: 自然语言解读
        top_n: 展示 Top N 因子

    Returns:
        PDF 文件的字节内容
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise ImportError("需要安装 reportlab：pip install reportlab")

    # 注册中文字体
    _register_chinese_font(pdfmetrics)

    # 创建 PDF 缓冲区
    buffer = io.BytesIO()

    # 创建文档
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    # 获取样式
    styles = getSampleStyleSheet()

    # 创建中文样式
    title_style = ParagraphStyle(
        'ChineseTitle',
        parent=styles['Title'],
        fontName='ChineseFont',
        fontSize=18,
        spaceAfter=20,
    )

    heading_style = ParagraphStyle(
        'ChineseHeading',
        parent=styles['Heading2'],
        fontName='ChineseFont',
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
    )

    body_style = ParagraphStyle(
        'ChineseBody',
        parent=styles['Normal'],
        fontName='ChineseFont',
        fontSize=10,
        spaceAfter=6,
    )

    # 构建内容
    story = []

    # 标题
    story.append(Paragraph(f"指标拆解分析报告：{config.name}", title_style))

    # 分析解读
    story.append(Paragraph("分析解读", heading_style))
    for line in summary.split('\n'):
        if line.strip():
            story.append(Paragraph(line, body_style))

    story.append(Spacer(1, 20))

    # 可视化图表
    story.append(Paragraph("可视化图表", heading_style))
    chart_images = _generate_chart_images(result, config, top_n)
    for img_data in chart_images:
        img = Image(io.BytesIO(img_data), width=16 * cm, height=10 * cm)
        story.append(img)
        story.append(Spacer(1, 10))

    story.append(Spacer(1, 20))

    # 拆解结果表格
    story.append(Paragraph("拆解结果", heading_style))
    table_data = [["因子", "贡献值", "贡献率(%)"]]
    for factor in result.contributions:
        table_data.append([
            factor.name,
            f"{factor.value_change:.2f}",
            f"{factor.contribution_rate * 100:.2f}%",
        ])

    table = Table(table_data, colWidths=[5 * cm, 4 * cm, 4 * cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ddd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
    ]))
    story.append(table)

    # 页脚信息
    story.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'Footer',
        parent=body_style,
        fontSize=8,
        textColor=colors.HexColor('#7f8c8d'),
    )
    story.append(Paragraph(
        f"拆解方法：{result.method.value} | MECE：{'是' if result.is_mece else '否'} | 因子数量：{len(result.contributions)}",
        footer_style
    ))
    story.append(Paragraph(f"报告生成时间：{_get_current_time()}", footer_style))

    # 生成 PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _register_chinese_font(pdfmetrics):
    """注册中文字体"""
    from reportlab.pdfbase.ttfonts import TTFont
    import platform
    system = platform.system()

    font_paths = []
    if system == "Darwin":  # macOS
        font_paths = [
            "/System/Library/Fonts/STHeiti Light.ttc",
            "/System/Library/Fonts/PingFang.ttc",
            "/System/Library/Fonts/Hiragino Sans GB.ttc",
        ]
    elif system == "Windows":
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
        ]
    else:  # Linux
        font_paths = [
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]

    for font_path in font_paths:
        try:
            pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
            return
        except Exception:
            continue

    # 如果没有找到中文字体，使用默认字体
    raise RuntimeError("未找到中文字体文件，请安装中文字体后重试")


def _generate_chart_images(result: DecompositionResult, config: MetricConfig, top_n: int) -> list:
    """生成图表图片"""
    images = []

    # 根据分析模式生成不同图表
    if result.mode == AnalysisMode.DYNAMIC:
        if result.method == DecompositionMethod.DUAL_FACTOR:
            fig = dual_factor_grouped_bar(result, top_n=top_n)
        else:
            fig = waterfall_chart(result, top_n=top_n)
    else:
        fig = composition_pie(result)

    # 转换为图片
    img_bytes = pio.to_image(fig, format="png", width=800, height=500)
    images.append(img_bytes)

    # 贡献率柱状图
    fig2 = contribution_bar(result, top_n=top_n)
    img_bytes2 = pio.to_image(fig2, format="png", width=800, height=500)
    images.append(img_bytes2)

    return images


def _get_current_time() -> str:
    """获取当前时间"""
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
